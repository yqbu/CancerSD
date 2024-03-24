import numpy as np
import pandas as pd
import xlrd
import openpyxl
import process_utils


if __name__ == '__main__':
    workbook = xlrd.open_workbook('../../Data/ACRG/patient_diagnosis_ACRG.xls')
    worksheet = workbook.sheet_by_index(0)

    content = []
    for row in range(worksheet.nrows):
        row_content = []
        for col in range(worksheet.ncols):
            cell_value = worksheet.cell(row, col).value
            row_content.append(cell_value)
        content.append(row_content)
    content = np.array(content)
    title = content[0, :]
    content = content[1:, :]
    clinical = pd.DataFrame(data=content, columns=title)
    clinical = clinical[['Tumor ID', 'Sample\nName', 'pStage']]
    clinical.rename(columns={'Sample\nName': 'sample', 'pStage': 'stage', 'Tumor ID': 'Sample ID'}, inplace=True)
    clinical['Sample ID'] = clinical['Sample ID'].astype(dtype=float)
    clinical['sample'] = clinical['sample'].str.strip()

    workbook = openpyxl.load_workbook('../../Data/ACRG/patient_diagnosis_match.xlsx')
    worksheet = workbook['Fig 3e survival data']
    subtype_map = {1: 'CIN', 2: 'MSI', 3: 'EBV', 4: 'GS'}

    content = []
    for row in worksheet.iter_rows(values_only=True):
        content.append(row)
    content = np.array(content)
    title = content[0, :]
    content = content[1:, :]

    subtypes = pd.DataFrame(data=content, columns=title)
    subtypes = subtypes[['Sample ID', 'Mol subtype']]
    subtypes['Mol subtype'] = subtypes['Mol subtype'].map(subtype_map)
    subtypes['Sample ID'] = subtypes['Sample ID'].astype(dtype=float)
    subtypes.rename(columns={'Mol subtype': 'subtype'}, inplace=True)

    diagnoses = clinical.merge(subtypes, on='Sample ID', how='inner').set_index('sample').T
    diagnoses.drop(index=['Sample ID'], inplace=True)
    diagnoses.index.rename('aspect', inplace=True)

    mRNA_df = pd.read_csv('../../Data/ACRG/expressions_mat.csv').T
    row_names = mRNA_df.index.values
    row_renames = dict()
    for name in row_names:
        split_name = name.split("_")
        if len(split_name) == 7:
            row_renames[name] = split_name[2]
        elif len(split_name) == 8:
            row_renames[name] = split_name[2] + '_' + split_name[3]
    mRNA_df.rename(index=row_renames, inplace=True)
    mRNA_df = mRNA_df.T
    mRNA_df.index.rename('mRNA', inplace=True)
    mRNA_df.to_csv('../../Data//ACRG/mRNA_original.csv')

    mRNA_df = pd.read_csv('../../../Data//ACRG/mRNA_original.csv').set_index('mRNA').T
    selected_molecules = list(np.load('../../../Data//Molecule_selected/mRNA_selected.npy', allow_pickle=True))
    mRNA_df = process_utils.normalize_omics_data('ACRG', 'mRNA', mRNA_df.T)

    mRNA_df = mRNA_df.loc[list(set(selected_molecules) & set(mRNA_df.index.values)), diagnoses.columns]
    mRNA_df = process_utils.omics_extend_as(mRNA_df, selected_molecules)
    mRNA_df.index.rename('mRNA', inplace=True)
    mRNA_df.to_csv('../../Data//ACRG/mRNA.csv', encoding='utf-8')
